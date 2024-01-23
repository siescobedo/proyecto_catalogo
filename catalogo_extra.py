import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import openpyxl

class BuscadorAccesosApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Buscador de Accesos")

        # Crear Frames para organizar la interfaz
        frame_top = ttk.Frame(root)
        frame_top.pack(padx=10, pady=10)

        frame_bottom = ttk.Frame(root)
        frame_bottom.pack(padx=10, pady=10)

        # Botones para cargar archivos
        self.organica_button = ttk.Button(frame_top, text="Seleccionar archivo de Orgánica Internos", command=self.cargar_organica)
        self.organica_button.grid(row=0, column=0, padx=5, pady=5)

        self.catalogo_button = ttk.Button(frame_top, text="Seleccionar archivo de Catálogo", command=self.cargar_catalogo)
        self.catalogo_button.grid(row=0, column=1, padx=5, pady=5)

        # Entrada para ingresar el RUT
        self.rut_label = ttk.Label(frame_bottom, text="Ingrese un RUT:")
        self.rut_label.grid(row=0, column=0, padx=5, pady=5)

        self.rut_entry = ttk.Entry(frame_bottom, font=("Helvetica", 12))
        self.rut_entry.grid(row=0, column=1, padx=5, pady=5)

        # Botón para iniciar el procesamiento
        self.procesar_button = ttk.Button(frame_bottom, text="Procesar", command=self.procesar)
        self.procesar_button.grid(row=1, column=0, columnspan=2, pady=10)

        self.organica_procesada = ""
        self.catalogo_procesado = ""


    def cargar_archivo(self, tipo):
        file_path = filedialog.askopenfilename(title=f"Seleccionar archivo {tipo.capitalize()}", filetypes=[(f"{tipo} files", f"*.{tipo}")])
        return file_path

    def cargar_organica(self):
        self.organica_path = self.cargar_archivo("xlsx")

    def cargar_catalogo(self):
        self.catalogo_path = self.cargar_archivo("xlsx")

    def procesar_organica(self, organica_path):
        wb_o = openpyxl.load_workbook(organica_path)
        ws_o = wb_o.active

        organica_c = {}
        for cell in ws_o[1]:
            if cell.value != None:
                organica_c[cell.value] = cell.column_letter

        self.rut_dict = {}
        for i in range(2,ws_o.max_row):
            rut = ws_o[f"{organica_c['RUT']}{i}"]
            ur = ws_o[f"{organica_c['UR']}{i}"]
            cargo = ws_o[f"{organica_c['Cargo']}{i}"]
            self.rut_dict[rut.value] = f"{ur.value}-{cargo.value}"
        
        self.organica_procesada = organica_path
    
    def make_dicts(self, dicts, ws, catalogo_c):
        [concat_dict, accesos] = dicts
        for i in range(2,ws.max_row):
            ur = ws[f"{catalogo_c['UR']}{i}"]
            cargo = ws[f"{catalogo_c['Cargo']}{i}"]
            rol = ws[f"{catalogo_c['Rol']}{i}"]
            app = ws[f"{catalogo_c['Aplicacion']}{i}"]
            perfil = ws[f"{catalogo_c['Perfil']}{i}"]

            concat = f"{ur.value}-{cargo.value}"
            if concat not in concat_dict.keys():
                concat_dict[concat] = set()
                concat_dict[concat].add(rol.value)
            else:
                concat_dict[concat].add(rol.value)

            if app.value!= None:
                if rol.value not in accesos.keys():
                    accesos[rol.value] = set()
                    accesos[rol.value].add((app.value, perfil.value))
                else:
                    accesos[rol.value].add((app.value, perfil.value))
        
        return [concat_dict, accesos]

    def procesar_catalogo(self, catalogo_path):
        wb_c = openpyxl.load_workbook(catalogo_path)
        ws_c_1 = wb_c['Sucursales']
        ws_c_2 = wb_c['Serv.Centrales']
        ws_c_3 = wb_c['Contac Center']

        catalogo_c = {}
        for cell in ws_c_1[1]:
            if cell.value != None:
                catalogo_c[cell.value] = cell.column_letter
        
        catalogo_dict = self.make_dicts(self.make_dicts(self.make_dicts([{},{}], ws_c_1, catalogo_c),ws_c_2, catalogo_c), ws_c_3, catalogo_c)

        self.concat_dict = catalogo_dict[0]
        self.accesos_dict = catalogo_dict[1]
        self.catalogo_procesado = catalogo_path

    def procesar(self):
        rut_ingresado = self.rut_entry.get()
        if self.organica_path is not None and self.catalogo_path is not None and rut_ingresado:
            if self.organica_path != self.organica_procesada:
                self.procesar_organica(self.organica_path)
            if self.catalogo_path != self.catalogo_procesado:
                self.procesar_catalogo(self.catalogo_path)


            roles = self.concat_dict.get(self.rut_dict.get(rut_ingresado, []), {})

            resultado_text = ""
            for rol in roles:
                resultado_text += f'Accesos del Rol: {rol}\n'
                accesos = self.accesos_dict[rol]
                for id, val in enumerate(accesos):
                    resultado_text += f"{id+1}. Aplicación: {val[0]}, Perfil: {val[1]}\n"

            # Mostrar resultado en una nueva ventana
            resultado_window = tk.Toplevel(self.root)
            resultado_label = tk.Label(resultado_window, text=resultado_text, padx=10, pady=10)
            resultado_label.pack()
        else:
            tk.messagebox.showwarning("Error", "Debe seleccionar ambos archivos y proporcionar un RUT.")

if __name__ == "__main__":
    root = tk.Tk()
    app = BuscadorAccesosApp(root)
    root.mainloop()