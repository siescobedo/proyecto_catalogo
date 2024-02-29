import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import scrolledtext
import openpyxl

class BuscadorAccesosApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Buscador de Accesos")
        self.root.geometry('350x250')

        self.root.configure(bg='#ECECEC')
        style = ttk.Style()
        style.theme_create('custom', settings={
            'TLabel': {'configure': {'background': '#EC0000', 'foreground': '#FFFFFF', 'font': ('Helvetica', 11)}},
            'TFrame': {'configure': {'background': '#EC0000', 'borderwidth': 0, 'relief': 'flat'}},
            'TButton': {'configure': {'background':'#DEEDF2', 'font': ('Helvetica', 11), 'anchor': 'center'}}}
            )

        style.theme_use('custom')

        frame = ttk.Frame(root)
        frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(frame, text="Buscador de Accesos", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10, sticky='en')

        ttk.Label(frame, text="Orgánica:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        ttk.Button(frame, text="Seleccionar archivo", command=self.cargar_organica, cursor="hand2").grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        
        ttk.Label(frame, text="Catálogo:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        ttk.Button(frame, text="Seleccionar archivo", command=self.cargar_catalogo, cursor="hand2").grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(frame, text="Ingrese RUT:").grid(row=3, column=0, padx=5, pady=5, sticky='e')
        self.rut_entry = ttk.Entry(frame, font=("Helvetica", 11))
        self.rut_entry.grid(row=3, column=1, padx=5, pady=5, sticky='ew')

        ttk.Button(frame, text="Buscar", command=self.procesar,cursor="hand2", width=20).grid(row=4, column=1, columnspan=2, pady=10, sticky='n')

        self.organica_procesada = ""
        self.catalogo_procesado = ""


    def cargar_archivo(self, tipo):
        file_path = filedialog.askopenfilename(title=f"Seleccionar archivo {tipo.capitalize()}", filetypes=[(f"{tipo} files", f"*.{tipo}")])
        return file_path

    def cargar_organica(self):
        self.organica_path = self.cargar_archivo("xlsx")

    def cargar_catalogo(self):
        self.catalogo_path = self.cargar_archivo("xlsx")

    def procesar_organica(self, organica_path):
        wb_o = openpyxl.load_workbook(organica_path)
        ws_o = wb_o.active

        organica_c = {}
        for cell in ws_o[1]:
            if cell.value != None:
                organica_c[cell.value] = cell.column_letter

        self.rut_dict = {}
        for i in range(2,ws_o.max_row+1):
            rut = ws_o[f"{organica_c['Rut']}{i}"]
            ur = ws_o[f"{organica_c['UR']}{i}"]
            cargo = ws_o[f"{organica_c['Cargo']}{i}"]
            self.rut_dict[rut.value] = f"{ur.value}-{cargo.value}"
        
        self.organica_procesada = organica_path
    
    def make_dicts(self, dicts, ws, catalogo_c):
        [concat_dict, accesos] = dicts
        for i in range(2,ws.max_row+1):
            ur = ws[f"{catalogo_c['CODIGOUR']}{i}"]
            cargo = ws[f"{catalogo_c['CODIGOCARGO']}{i}"]
            rol = ws[f"{catalogo_c['ROL']}{i}"]
            app = ws[f"{catalogo_c['APLICACION']}{i}"]
            perfil = ws[f"{catalogo_c['PERFIL']}{i}"]
            seccion = ws[f"{catalogo_c['SECCION']}{i}"]

            concat = f"{ur.value}-{cargo.value}"
            if concat not in concat_dict.keys():
                concat_dict[concat] = set()
                concat_dict[concat].add(f"{rol.value}; {seccion.value}")
            else:
                concat_dict[concat].add(f"{rol.value}; {seccion.value}")

            concat2 = f"{rol.value}; {seccion.value}"
            if app.value!= None:
                if concat2 not in accesos.keys():
                    accesos[concat2] = set()
                    accesos[concat2].add((app.value, perfil.value))
                else:
                    accesos[concat2].add((app.value, perfil.value))
        
        return [concat_dict, accesos]

    def procesar_catalogo(self, catalogo_path):
        wb_c = openpyxl.load_workbook(catalogo_path)
        ws_c = wb_c.active

        catalogo_c = {}
        for cell in ws_c[1]:
            if cell.value != None:
                catalogo_c[cell.value] = cell.column_letter
        
        catalogo_dict = self.make_dicts([{},{}], ws_c, catalogo_c)

        self.concat_dict = catalogo_dict[0]
        self.accesos_dict = catalogo_dict[1]
        self.catalogo_procesado = catalogo_path

    def mostrar_resultados(self, resultado_text):
        # Mostrar resultado en una nueva ventana
        resultado_window = tk.Toplevel(self.root)
        resultado_window.title("Resultados")

        resultado_window.configure(bg='#ECECEC')

        style = ttk.Style()
        style.theme_use('custom')

        frame = ttk.Frame(resultado_window)
        frame.pack(expand=1, fill="both", padx=25, pady=25)

        ttk.Label(frame, text=f"Accesos {self.rut}", font=("Helvetica", 16)).pack(pady=20)

        scroll_text = scrolledtext.ScrolledText(frame, wrap=tk.WORD, width=40, height=10, font=("Helvetica", 11))
        scroll_text.insert(tk.END, resultado_text)
        scroll_text.config(padx=10, pady=10)
        scroll_text.pack(expand=True, fill="both", padx=15, pady=15)

        resultado_window.protocol("WM_DELETE_WINDOW", resultado_window.destroy)  # Maneja el cierre de la ventana


    def procesar(self):
        self.rut = self.rut_entry.get()
        if self.organica_path is not None and self.catalogo_path is not None and self.rut:
            if self.organica_path != self.organica_procesada:
                self.procesar_organica(self.organica_path)
            if self.catalogo_path != self.catalogo_procesado:
                self.procesar_catalogo(self.catalogo_path)


            roles = self.concat_dict.get(self.rut_dict.get(self.rut, []), {})

            resultado_text = ""
            for rol in roles:
                resultado_text += f'Accesos del Rol: {rol}\n'
                accesos = self.accesos_dict[rol]
                for id, val in enumerate(accesos):
                    resultado_text += f"{id+1}. Aplicación: {val[0]}, Perfil: {val[1]}\n"
                resultado_text += f"\n\n"
            self.mostrar_resultados(resultado_text)
        else:
            tk.messagebox.showwarning("Error", "Debe seleccionar ambos archivos y proporcionar un RUT.")

if __name__ == "__main__":
    root = tk.Tk()
    app = BuscadorAccesosApp(root)
    root.mainloop()