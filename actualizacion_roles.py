import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import scrolledtext
import openpyxl
import os
import sys
from openpyxl.utils import column_index_from_string


class ActualizadorRolesApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Actualizador de roles")
        self.root.geometry('350x250')

        self.root.configure(bg='#ECECEC')
        style = ttk.Style()
        style.theme_create('custom', settings={
            'TLabel': {'configure': {'background': '#EC0000', 'foreground': '#FFFFFF', 'font': ('Helvetica', 11)}},
            'TFrame': {'configure': {'background': '#EC0000', 'borderwidth': 0, 'relief': 'flat'}},
            'TButton': {'configure': {'background':'#DEEDF2', 'font': ('Helvetica', 11), 'anchor': 'center'}}}
            )

        style.theme_use('custom')

        frame = ttk.Frame(root)
        frame.pack(expand=True, fill="both", padx=20, pady=20)

        ttk.Label(frame, text="Actualizador de roles", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10, sticky='en')

        ttk.Label(frame, text="Orgánica:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        ttk.Button(frame, text="Seleccionar archivo", command=self.cargar_organica, cursor="hand2").grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        
        ttk.Label(frame, text="Catálogo:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
        ttk.Button(frame, text="Seleccionar archivo", command=self.cargar_catalogo, cursor="hand2").grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        self.progreso_label = ttk.Label(frame, text="", font=("Helvetica", 12))
        self.progreso_label.grid(row=3, column=0, columnspan=2, pady=10)

        ttk.Button(frame, text="Actualizar", command=self.procesar,cursor="hand2", width=20).grid(row=4, column=1, columnspan=2, pady=10, sticky='n')

        self.organica_procesada = ""
        self.catalogo_procesado = ""


    def cargar_archivo(self, tipo):
        file_path = filedialog.askopenfilename(title=f"Seleccionar archivo {tipo.capitalize()}", filetypes=[(f"{tipo} files", f"*.{tipo}")])
        return file_path

    def cargar_organica(self):
        self.organica_path = self.cargar_archivo("xlsx")

    def cargar_catalogo(self):
        self.catalogo_path = self.cargar_archivo("xlsx")
    
    def actualizar_progreso(self, mensaje):
        self.progreso_label.config(text=mensaje)

    def procesar_organica(self, organica_path):
        self.actualizar_progreso("Procesando orgánica...")
        wb_o = openpyxl.load_workbook(organica_path)
        ws_o = wb_o.active

        organica_c = {}
        for cell in ws_o[1]:
            if cell.value != None:
                organica_c[cell.value] = cell.column_letter

        self.organica_concats = set()
        self.concats_dict = {}

        for i in range(2,ws_o.max_row+1):
            ur = ws_o[f"{organica_c['UR']}{i}"]
            cargo = ws_o[f"{organica_c['Cargo']}{i}"]
            gls_cargo = ws_o[f"{organica_c['GlsCargo']}{i}"]
            gls_ur = ws_o[f"{organica_c['GlsUR']}{i}"]
            concat = f"{ur.value}-{cargo.value}"
            if concat in self.organica_concats:
                agregar = True
                for i in self.concats_dict[concat]:
                    if i == {'CODIGOCARGO': cargo.value, 'CARGO': gls_cargo.value, 'CODIGOUR': ur.value, 'UNIRELUR': gls_ur.value}:
                        agregar = False
                
                if agregar:
                    self.concats_dict[concat].append({'CODIGOCARGO': cargo.value, 'CARGO': gls_cargo.value, 'CODIGOUR': ur.value, 'UNIRELUR': gls_ur.value})
            else:
                self.organica_concats.add(concat)
                self.concats_dict[concat] = [{'CODIGOCARGO': cargo.value, 'CARGO': gls_cargo.value, 'CODIGOUR': ur.value, 'UNIRELUR': gls_ur.value}]

        self.organica_procesada = organica_path

    def procesar_catalogo(self, catalogo_path):
        self.actualizar_progreso("Procesando catálogo...")
        wb_c = openpyxl.load_workbook(catalogo_path)
        self.ws_c = wb_c.active

        self.catalogo_c = {}
        for cell in self.ws_c[1]:
            if cell.value != None:
                self.catalogo_c[cell.value] = cell.column_letter

        self.catalogo_concats = set()

        for i in range(2,self.ws_c.max_row+1):
            ur = self.ws_c[f"{self.catalogo_c['CODIGOUR']}{i}"]
            cargo = self.ws_c[f"{self.catalogo_c['CODIGOCARGO']}{i}"]
            self.catalogo_concats.add(f"{ur.value}-{cargo.value}")
        
        self.catalogo_procesado = catalogo_path

    def crear_catalogo(self, roles):
        self.actualizar_progreso("Creando nuevo archivo...")
        
        # copiar todo a archivo nuevo
            
        wb_nuevo = openpyxl.Workbook()
        hoja_nueva = wb_nuevo.active
        hoja_nueva.title = "Informe"

        for row in self.ws_c.iter_rows(values_only=True):
            hoja_nueva.append(row)

        # agregar filas 
        
        self.actualizar_progreso("Agregando nuevas concatenaciones...")
        for concatenacion_key in roles:
            concatenacion_contenido = self.concats_dict[concatenacion_key]

            nueva_fila = [None] * len(self.catalogo_c)
            
            for concatenacion in concatenacion_contenido:
                for col_name in concatenacion.keys():
                    if col_name in self.catalogo_c.keys():
                        col_letter = self.catalogo_c[col_name]
                        col_index = column_index_from_string(col_letter)
                        nueva_fila[col_index - 1] = concatenacion[col_name]
            
                hoja_nueva.append(nueva_fila)

        self.actualizar_progreso("Guardando nuevo archivo...")
        # guardar nuevo archivo
        directorio_actual = os.path.dirname(sys.argv[0])
        archivo_nuevo = os.path.join(directorio_actual, "catalogo_actualizado.xlsx")
        wb_nuevo.save(archivo_nuevo)


    def procesar(self):
        if self.organica_path is not None and self.catalogo_path is not None:
            if self.organica_path != self.organica_procesada:
                self.procesar_organica(self.organica_path)
            if self.catalogo_path != self.catalogo_procesado:
                self.procesar_catalogo(self.catalogo_path)

            roles = self.organica_concats - self.catalogo_concats
            
            self.crear_catalogo(list(roles))
            self.actualizar_progreso("¡Actualización y archivo nuevo terminados!")
        else:
            tk.messagebox.showwarning("Error", "Debe seleccionar ambos archivos.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ActualizadorRolesApp(root)
    root.mainloop()