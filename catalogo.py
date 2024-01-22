import openpyxl
import tkinter as tk
from tkinter import filedialog

def cargar_archivo(tipo):
    file_path = filedialog.askopenfilename(title=f"Seleccionar archivo {tipo.capitalize()}", filetypes=[(f"{tipo} files", f"*.{tipo}")])
    return file_path

def busqueda_de_accesos(rut, rut_dict, catalogo_dict):
    roles = catalogo_dict[rut_dict[rut]]
    for rol in roles:
        print(f'Accesos del Rol: {rol}')
        accesos = roles[rol]
        for id, val in enumerate(accesos):
            print(f"{id+1}. Aplicación: {val['Aplicacion']}, Perfil: {val['Perfil']}")

def procesar_organica(organica_path):
    wb_o = openpyxl.load_workbook(organica_path)
    ws_o = wb_o.active

    organica_c = {}
    for cell in ws_o[1]:
        if cell.value != None:
            organica_c[cell.value] = cell.column_letter

    rut_dict = {}
    for i in range(2,ws_o.max_row):
        rut = ws_o[f"{organica_c['RUT']}{i}"]
        ur = ws_o[f"{organica_c['UR']}{i}"]
        cargo = ws_o[f"{organica_c['Cargo']}{i}"]
        rut_dict[rut.value] = f"{ur.value}-{cargo.value}"
    
    return rut_dict

def make_concat_dict(concat_dict, ws, catalogo_c):
  for i in range(2,ws.max_row):
    ur = ws[f"{catalogo_c['UR']}{i}"]
    cargo = ws[f"{catalogo_c['Cargo']}{i}"]
    rol = ws[f"{catalogo_c['Rol']}{i}"]
    app = ws[f"{catalogo_c['Aplicacion']}{i}"]
    perfil = ws[f"{catalogo_c['Perfil']}{i}"]

    concat = f"{ur.value}-{cargo.value}"
    if concat not in concat_dict.keys():
      concat_dict[concat] = {rol.value: [{'Aplicacion': app.value, 'Perfil': perfil.value}]}
    else:
      if rol.value in concat_dict[concat].keys():
        concat_dict[concat][rol.value].append({'Aplicacion': app.value, 'Perfil': perfil.value})
      else:
        concat_dict[concat][rol.value] = [{'Aplicacion': app.value, 'Perfil': perfil.value}]

  return concat_dict

def procesar_catalogo(catalogo_path):
    wb_c = openpyxl.load_workbook(catalogo_path)
    ws_c_1 = wb_c['Sucursales']
    ws_c_2 = wb_c['Serv.Centrales']
    ws_c_3 = wb_c['Contac Center']

    catalogo_c = {}
    for cell in ws_c_1[1]:
        if cell.value != None:
            catalogo_c[cell.value] = cell.column_letter
    
    catalogo_dict = make_concat_dict(make_concat_dict(make_concat_dict({}, ws_c_1, catalogo_c),ws_c_2, catalogo_c), ws_c_3, catalogo_c)

    return catalogo_dict

def procesar_archivos(organica_path, catalogo_path):
    rut_dict = procesar_organica(organica_path)
    catalogo_dict = procesar_catalogo(catalogo_path)

    # Aquí puedes construir tu interfaz gráfica para que el usuario ingrese el RUT
    rut_ingresado = input("Ingrese un RUT: ")
    busqueda_de_accesos(rut_ingresado, rut_dict, catalogo_dict)

def main():
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal de tkinter

    print("Selecciona el archivo de orgánica internos")
    organica_path = cargar_archivo("xlsx")

    print("Selecciona el archivo de catálogo")
    catalogo_path = cargar_archivo("xlsx")

    procesar_archivos(organica_path, catalogo_path)

if __name__ == "__main__":
    main()